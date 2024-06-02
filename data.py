import glob
import json
import re
from post import scrape_post
import pandas as pd
from openpyxl import load_workbook
from selenium.webdriver.chrome.webdriver import WebDriver

def scrape_from_links(browser: WebDriver, batch_size: int):
    pattern = 'links_*.json'
    file_list = sorted(glob.glob(pattern), key=lambda x: int(re.search(r'\d+', x).group()))
    all_data = []

    for file in file_list:
        with open(file, 'r') as file:
            data = json.load(file)
            all_data.extend(data)

    post_count = 0
    posts = []

    for link in all_data:
        data = scrape_post(browser, link)
        posts.append(link)
        post_count += 1

        if post_count % batch_size == 0:
            print(f"Saving batch of {batch_size} posts...")
            with open(f"posts_{post_count}.json", "w") as file:
                json.dump(posts, file, indent=4)
            posts.clear()

def posts_to_excel():
    pattern = 'posts_*.json'
    file_list = sorted(glob.glob(pattern), key=lambda x: int(re.search(r'\d+', x).group()))
    all_data = []

    for file in file_list:
        with open(file, 'r') as file:
            data = json.load(file)
            all_data.extend(data)

    df = pd.DataFrame((all_data))
    output_file = 'output.xlsx'
    df.to_excel(output_file, index=False)

    # Load the workbook and select the active worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Iterate over the cells in the 'url' column and add hyperlinks
    for cell in ws['I'][1:]:  # Assuming 'url' is in column D
        cell.hyperlink = cell.value
        cell.style = "Hyperlink"

    # Save the workbook
    wb.save(output_file)

    print(f"Data successfully written to {output_file}")